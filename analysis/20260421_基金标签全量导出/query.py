"""
基金标签全量导出（按基金类型输出独立 xlsx）

输出 4 个文件：
  - 基金标签_主动权益_{date}.xlsx  (001, 仅 c_type2_code=001001)
  - 基金标签_固收+_{date}.xlsx     (002)
  - 基金标签_债券_{date}.xlsx      (003)
  - 基金标签_混合_{date}.xlsx      (004)
"""
from pathlib import Path
import pandas as pd
from utils.db_connector import DorisConnector
from utils.log import setup_logger

logger = setup_logger(__name__)
ENV = 'dev'

REPORT_DATE = '2025-12-31'   # ← 修改这里切换报告期

OUT_DIR = Path(__file__).parent / 'data'
OUT_DIR.mkdir(exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# 公共规模子查询（内联使用）
# ─────────────────────────────────────────────────────────────────────────────
_SCALE_SUB = """
LEFT JOIN (
    SELECT c_fd_code, c_report_date, MAX(c_fund_nav_total) AS c_fund_nav_total
    FROM tb_fd_asset_allocation
    WHERE c_style IN ('01', '03', '05', '06')
    GROUP BY c_fd_code, c_report_date
) sc ON sc.c_fd_code = cat.c_fd_code AND sc.c_report_date = cat.c_report_date"""


# ─────────────────────────────────────────────────────────────────────────────
# SQL 定义
# ─────────────────────────────────────────────────────────────────────────────

_SQL_EQUITY = """
SELECT
    b.c_fd_code,
    b.c_short_name,
    b.c_company_name,
    b.c_class3_name,
    b.c_manager_name,
    ROUND(sc.c_fund_nav_total / 100000000, 2) AS c_scale_bn,
    -- 资产配置
    eq.c_stk_pos_level,
    eq.c_stk_pos_avg,
    eq.c_stk_timing,
    -- 区域 & 板块
    rs.c_region_tag,
    rs.c_sector_tag,
    rs.c_sector_pref,
    rs.c_sector_cycle,
    rs.c_sector_mfg,
    rs.c_sector_tech,
    rs.c_sector_consumer,
    rs.c_sector_pharma,
    rs.c_sector_fin,
    rs.c_sector_chg,
    -- 风格
    sty.c_size_tag,
    sty.c_vg_tag,
    sty.c_momentum_tag,
    sty.c_profit_tag,
    sty.c_quality_tag,
    -- 组合构建
    pf.c_sector_concent_tag,
    pf.c_ind_concent_tag,
    pf.c_stk_concent_tag,
    -- 主动管理
    pf.c_active_tag,
    pf.c_new_stk_tag,
    pf.c_crowd_tag,
    -- 交易特征
    pf.c_turnover_tag,
    pf.c_heavy_trade_tag,
    pf.c_buy_timing_tag,
    pf.c_sell_timing_tag
FROM tb_fd_category cat
JOIN tb_fd_basic_info b
    ON b.c_fd_code = cat.c_fd_code
    AND (b.c_init_code = b.c_fd_code OR b.c_init_code IS NULL)
LEFT JOIN (
    SELECT c_fd_code, c_report_date, MAX(c_fund_nav_total) AS c_fund_nav_total
    FROM tb_fd_asset_allocation
    WHERE c_style IN ('01', '03', '05', '06')
    GROUP BY c_fd_code, c_report_date
) sc ON sc.c_fd_code = cat.c_fd_code AND sc.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_asset_eq eq
    ON eq.c_fd_code = cat.c_fd_code AND eq.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_region_sector rs
    ON rs.c_fd_code = cat.c_fd_code AND rs.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_style sty
    ON sty.c_fd_code = cat.c_fd_code AND sty.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_portfolio pf
    ON pf.c_fd_code = cat.c_fd_code AND pf.c_report_date = cat.c_report_date
WHERE cat.c_type1_code = '001'
  AND cat.c_type2_code = '001001'
  AND cat.c_report_date = :report_date
ORDER BY b.c_fd_code ASC
"""

_SQL_FI_PLUS = """
SELECT
    b.c_fd_code,
    b.c_short_name,
    b.c_company_name,
    b.c_class3_name,
    b.c_manager_name,
    ROUND(sc.c_fund_nav_total / 100000000, 2) AS c_scale_bn,
    -- 资产配置
    fi.c_eq_risk_level,
    fi.c_stk_cb_strategy,
    fi.c_stk_timing,
    fi.c_cb_timing,
    fi.c_eq_pos_avg,
    fi.c_stk_pos_avg,
    fi.c_cb_pos_avg,
    -- 债券风格
    bd.c_bd_type_style,
    bd.c_bd_type_timing,
    bd.c_credit_pref,
    bd.c_credit_timing,
    bd.c_duration_pref,
    bd.c_duration_avg,
    bd.c_leverage_pref,
    bd.c_leverage_avg,
    -- 转债风格（转债仓位≥1%的基金才有值）
    cb.c_cb_concent_tag,
    cb.c_cb_ind_concent_tag,
    cb.c_cb_sector_tag,
    cb.c_cb_sector_pref,
    cb.c_cb_trade_tag,
    cb.c_cb_attr_tag,
    cb.c_cb_equity_tag,
    cb.c_cb_bond_tag,
    cb.c_cb_balance_tag,
    cb.c_stk_mktcap_tag,
    cb.c_stk_style_tag,
    -- 权益持仓标签（有股票仓位的才有值）
    rs.c_region_tag,
    rs.c_sector_tag    AS c_stk_sector_tag,
    rs.c_sector_pref   AS c_stk_sector_pref,
    sty.c_size_tag,
    sty.c_vg_tag,
    -- 绩效风险（近1年，period_code='03'）
    p1y.c_ann_ret  AS c_ann_ret_1y,
    p1y.c_ann_vol  AS c_ann_vol_1y,
    p1y.c_mdd      AS c_mdd_1y,
    p1y.c_sharpe   AS c_sharpe_1y,
    -- 绩效风险（近3年，period_code='05'；成立不足3年则为空）
    p3y.c_ann_ret  AS c_ann_ret_3y,
    p3y.c_ann_vol  AS c_ann_vol_3y,
    p3y.c_mdd      AS c_mdd_3y,
    p3y.c_sharpe   AS c_sharpe_3y
FROM tb_fd_category cat
JOIN tb_fd_basic_info b
    ON b.c_fd_code = cat.c_fd_code
    AND (b.c_init_code = b.c_fd_code OR b.c_init_code IS NULL)
LEFT JOIN (
    SELECT c_fd_code, c_report_date, MAX(c_fund_nav_total) AS c_fund_nav_total
    FROM tb_fd_asset_allocation
    WHERE c_style IN ('01', '03', '05', '06')
    GROUP BY c_fd_code, c_report_date
) sc ON sc.c_fd_code = cat.c_fd_code AND sc.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_asset_fi fi
    ON fi.c_fd_code = cat.c_fd_code AND fi.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_bd_style bd
    ON bd.c_fd_code = cat.c_fd_code AND bd.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_cb_style cb
    ON cb.c_fd_code = cat.c_fd_code AND cb.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_region_sector rs
    ON rs.c_fd_code = cat.c_fd_code AND rs.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_style sty
    ON sty.c_fd_code = cat.c_fd_code AND sty.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_perform_abs p1y
    ON p1y.c_fd_code = cat.c_fd_code
   AND p1y.c_trade_date = cat.c_report_date
   AND p1y.c_period_code = '03'
LEFT JOIN tb_fd_perform_abs p3y
    ON p3y.c_fd_code = cat.c_fd_code
   AND p3y.c_trade_date = cat.c_report_date
   AND p3y.c_period_code = '05'
WHERE cat.c_type1_code = '002'
  AND cat.c_report_date = :report_date
ORDER BY b.c_fd_code ASC
"""

_SQL_BOND = """
SELECT
    b.c_fd_code,
    b.c_short_name,
    b.c_company_name,
    b.c_class3_name,
    b.c_manager_name,
    ROUND(sc.c_fund_nav_total / 100000000, 2) AS c_scale_bn,
    -- 债券风格
    bd.c_bd_type_style,
    bd.c_bd_type_timing,
    bd.c_credit_pref,
    bd.c_credit_timing,
    bd.c_duration_pref,
    bd.c_duration_avg,
    bd.c_duration_chg_avg,
    bd.c_leverage_pref,
    bd.c_leverage_avg,
    bd.c_leverage_chg_avg,
    bd.c_is_locked
FROM tb_fd_category cat
JOIN tb_fd_basic_info b
    ON b.c_fd_code = cat.c_fd_code
    AND (b.c_init_code = b.c_fd_code OR b.c_init_code IS NULL)
LEFT JOIN (
    SELECT c_fd_code, c_report_date, MAX(c_fund_nav_total) AS c_fund_nav_total
    FROM tb_fd_asset_allocation
    WHERE c_style IN ('01', '03', '05', '06')
    GROUP BY c_fd_code, c_report_date
) sc ON sc.c_fd_code = cat.c_fd_code AND sc.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_bd_style bd
    ON bd.c_fd_code = cat.c_fd_code AND bd.c_report_date = cat.c_report_date
WHERE cat.c_type1_code = '003'
  AND cat.c_report_date = :report_date
ORDER BY b.c_fd_code ASC
"""

_SQL_MIX = """
SELECT
    b.c_fd_code,
    b.c_short_name,
    b.c_company_name,
    b.c_class3_name,
    b.c_manager_name,
    ROUND(sc.c_fund_nav_total / 100000000, 2) AS c_scale_bn,
    -- 资产配置
    mix.c_stk_bd_pref,
    mix.c_eq_strategy,
    mix.c_eq_timing,
    mix.c_stk_pos_avg,
    mix.c_cb_pos_avg,
    mix.c_bd_pos_avg,
    -- 区域 & 板块（有股票仓位才有值）
    rs.c_region_tag,
    rs.c_sector_tag,
    rs.c_sector_pref,
    rs.c_sector_cycle,
    rs.c_sector_mfg,
    rs.c_sector_tech,
    rs.c_sector_consumer,
    rs.c_sector_pharma,
    rs.c_sector_fin,
    -- 风格
    sty.c_size_tag,
    sty.c_vg_tag,
    sty.c_momentum_tag,
    -- 债券风格（有债券仓位才有值）
    bd.c_bd_type_style,
    bd.c_credit_pref,
    bd.c_duration_pref,
    bd.c_leverage_pref,
    -- 组合 & 交易
    pf.c_turnover_tag,
    pf.c_active_tag,
    pf.c_stk_concent_tag,
    pf.c_crowd_tag
FROM tb_fd_category cat
JOIN tb_fd_basic_info b
    ON b.c_fd_code = cat.c_fd_code
    AND (b.c_init_code = b.c_fd_code OR b.c_init_code IS NULL)
LEFT JOIN (
    SELECT c_fd_code, c_report_date, MAX(c_fund_nav_total) AS c_fund_nav_total
    FROM tb_fd_asset_allocation
    WHERE c_style IN ('01', '03', '05', '06')
    GROUP BY c_fd_code, c_report_date
) sc ON sc.c_fd_code = cat.c_fd_code AND sc.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_asset_mix mix
    ON mix.c_fd_code = cat.c_fd_code AND mix.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_region_sector rs
    ON rs.c_fd_code = cat.c_fd_code AND rs.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_style sty
    ON sty.c_fd_code = cat.c_fd_code AND sty.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_bd_style bd
    ON bd.c_fd_code = cat.c_fd_code AND bd.c_report_date = cat.c_report_date
LEFT JOIN tb_fd_tag_stk_portfolio pf
    ON pf.c_fd_code = cat.c_fd_code AND pf.c_report_date = cat.c_report_date
WHERE cat.c_type1_code = '004'
  AND cat.c_report_date = :report_date
ORDER BY b.c_fd_code ASC
"""


# ─────────────────────────────────────────────────────────────────────────────
# 列名映射（原始字段 → 中文显示名）
# ─────────────────────────────────────────────────────────────────────────────

_COMMON_COLS = {
    'c_fd_code':      '基金代码',
    'c_short_name':   '基金简称',
    'c_company_name': '基金公司',
    'c_class3_name':  '投资类型',
    'c_manager_name': '基金经理',
    'c_scale_bn':     '规模(亿元)',
}

_EQUITY_COLS = {
    **_COMMON_COLS,
    'c_stk_pos_level':     '仓位中枢',
    'c_stk_pos_avg':       '股票仓位均值(%)',
    'c_stk_timing':        '是否择时',
    'c_region_tag':        '区域特征',
    'c_sector_tag':        '板块标签',
    'c_sector_pref':       '板块偏好',
    'c_sector_cycle':      '周期(%)',
    'c_sector_mfg':        '中游制造(%)',
    'c_sector_tech':       '科技(%)',
    'c_sector_consumer':   '消费(%)',
    'c_sector_pharma':     '医药(%)',
    'c_sector_fin':        '金融地产(%)',
    'c_sector_chg':        '板块轮动度',
    'c_size_tag':          '市值标签',
    'c_vg_tag':            '价值-成长',
    'c_momentum_tag':      '动量标签',
    'c_profit_tag':        '盈利标签',
    'c_quality_tag':       '质量标签',
    'c_sector_concent_tag':'板块集中度',
    'c_ind_concent_tag':   '行业集中度',
    'c_stk_concent_tag':   '个股集中度',
    'c_active_tag':        '主动配置',
    'c_new_stk_tag':       '持股扩新',
    'c_crowd_tag':         '持股抱团',
    'c_turnover_tag':      '换手特征',
    'c_heavy_trade_tag':   '重仓持股',
    'c_buy_timing_tag':    '买入特征',
    'c_sell_timing_tag':   '卖出特征',
}

_FI_PLUS_COLS = {
    **_COMMON_COLS,
    'c_eq_risk_level':       '风险特征',
    'c_stk_cb_strategy':     '股票转债策略',
    'c_stk_timing':          '股票择时',
    'c_cb_timing':           '转债择时',
    'c_eq_pos_avg':          '权益仓位均值(%)',
    'c_stk_pos_avg':         '股票仓位均值(%)',
    'c_cb_pos_avg':          '转债仓位均值(%)',
    'c_bd_type_style':       '券种风格',
    'c_bd_type_timing':      '券种择时',
    'c_credit_pref':         '信用偏好',
    'c_credit_timing':       '信用择时',
    'c_duration_pref':       '久期偏好',
    'c_duration_avg':        '久期均值(年)',
    'c_leverage_pref':       '杠杆偏好',
    'c_leverage_avg':        '杠杆率均值',
    'c_cb_concent_tag':      'CB个券集中度',
    'c_cb_ind_concent_tag':  'CB行业集中度',
    'c_cb_sector_tag':       'CB板块标签',
    'c_cb_sector_pref':      'CB板块偏好',
    'c_cb_trade_tag':        'CB交易特征',
    'c_cb_attr_tag':         'CB属性(股/债)',
    'c_cb_equity_tag':       'CB股性',
    'c_cb_bond_tag':         'CB债性',
    'c_cb_balance_tag':      'CB余额',
    'c_stk_mktcap_tag':      '正股市值',
    'c_stk_style_tag':       '正股风格',
    'c_region_tag':          '股票区域',
    'c_stk_sector_tag':      '股票板块',
    'c_stk_sector_pref':     '股票板块偏好',
    'c_size_tag':            '股票市值风格',
    'c_vg_tag':              '股票价值-成长',
    # 绩效风险标签
    'c_ann_ret_1y':          '近1年年化收益率(%)',
    'c_ann_vol_1y':          '近1年年化波动率(%)',
    'c_mdd_1y':              '近1年最大回撤(%)',
    'c_sharpe_1y':           '近1年夏普比率',
    'c_ann_ret_3y':          '近3年年化收益率(%)',
    'c_ann_vol_3y':          '近3年年化波动率(%)',
    'c_mdd_3y':              '近3年最大回撤(%)',
    'c_sharpe_3y':           '近3年夏普比率',
}

_BOND_COLS = {
    **_COMMON_COLS,
    'c_bd_type_style':    '券种风格',
    'c_bd_type_timing':   '券种择时',
    'c_credit_pref':      '信用偏好',
    'c_credit_timing':    '信用择时',
    'c_duration_pref':    '久期偏好',
    'c_duration_avg':     '久期均值(年)',
    'c_duration_chg_avg': '久期变动均值(年)',
    'c_leverage_pref':    '杠杆偏好',
    'c_leverage_avg':     '杠杆率均值',
    'c_leverage_chg_avg': '杠杆率变动均值',
    'c_is_locked':        '是否定开',
}

_MIX_COLS = {
    **_COMMON_COLS,
    'c_stk_bd_pref':    '股债偏好',
    'c_eq_strategy':    '权益策略',
    'c_eq_timing':      '权益择时',
    'c_stk_pos_avg':    '股票仓位均值(%)',
    'c_cb_pos_avg':     '转债仓位均值(%)',
    'c_bd_pos_avg':     '债券仓位均值(%)',
    'c_region_tag':     '区域特征',
    'c_sector_tag':     '板块标签',
    'c_sector_pref':    '板块偏好',
    'c_sector_cycle':   '周期(%)',
    'c_sector_mfg':     '中游制造(%)',
    'c_sector_tech':    '科技(%)',
    'c_sector_consumer':'消费(%)',
    'c_sector_pharma':  '医药(%)',
    'c_sector_fin':     '金融地产(%)',
    'c_size_tag':       '市值标签',
    'c_vg_tag':         '价值-成长',
    'c_momentum_tag':   '动量标签',
    'c_bd_type_style':  '券种风格',
    'c_credit_pref':    '信用偏好',
    'c_duration_pref':  '久期偏好',
    'c_leverage_pref':  '杠杆偏好',
    'c_turnover_tag':   '换手特征',
    'c_active_tag':     '主动配置',
    'c_stk_concent_tag':'个股集中度',
    'c_crowd_tag':      '持股抱团',
}


# ─────────────────────────────────────────────────────────────────────────────
# 固收+ 双行表头配置：{分类名: 列数}，顺序与 _FI_PLUS_COLS 字段顺序一致
# ─────────────────────────────────────────────────────────────────────────────
_FI_PLUS_CATEGORY = {
    '基本信息':    6,
    '资产配置标签': 7,
    '债券风格标签': 8,
    '转债风格标签': 11,
    '权益风格标签': 5,
    '绩效风险标签': 8,
}


# ─────────────────────────────────────────────────────────────────────────────
# 查询函数
# ─────────────────────────────────────────────────────────────────────────────

def _fetch(doris: DorisConnector, sql: str, col_map: dict) -> pd.DataFrame:
    df = doris.query(sql, report_date=REPORT_DATE)
    existing = [c for c in col_map if c in df.columns]
    return df[existing].rename(columns=col_map)


def _save(df: pd.DataFrame, filename: str) -> Path:
    path = OUT_DIR / filename
    df.to_excel(path, index=False)
    logger.info(f"已保存 {filename}  ({len(df)} 行)")
    return path


def _save_fi_plus(df: pd.DataFrame, filename: str) -> Path:
    """固收+ 专用：双行表头（第1行分类合并，第2行字段名），楷体11pt。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = OUT_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = '固收+标签'

    bold_font  = Font(name='楷体', size=11, bold=True)
    data_font  = Font(name='楷体', size=11)
    cat_fill   = PatternFill('solid', fgColor='D9D9D9')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center')

    # ── 第1行：分类（合并单元格）───────────────────────────────────────────
    col = 1
    for cat_name, n_cols in _FI_PLUS_CATEGORY.items():
        start_col = col
        end_col   = col + n_cols - 1
        if n_cols > 1:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1,   end_column=end_col,
            )
        cell = ws.cell(row=1, column=start_col, value=cat_name)
        cell.font      = bold_font
        cell.fill      = cat_fill
        cell.alignment = center_align
        col += n_cols

    # ── 第2行：字段名 ──────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=c_idx, value=col_name)
        cell.font      = bold_font
        cell.alignment = center_align

    # ── 第3行起：数据 ─────────────────────────────────────────────────────
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = data_font
            cell.alignment = left_align
            if isinstance(value, float):
                cell.number_format = '0.00'

    # ── 冻结前两行 ─────────────────────────────────────────────────────────
    ws.freeze_panes = 'A3'

    wb.save(path)
    logger.info(f"已保存 {filename}  ({len(df)} 行，双行表头)")
    return path


def run_fi_plus():
    """仅导出固收+ 标签（双行表头，含绩效风险）。"""
    with DorisConnector(ENV) as doris:
        logger.info(f"查询固收+ 报告期 {REPORT_DATE}")
        df = _fetch(doris, _SQL_FI_PLUS, _FI_PLUS_COLS)
    logger.info(f"固收+ 行数: {len(df)}")
    _save_fi_plus(df, f'基金标签_固收+_{REPORT_DATE}.xlsx')


def run():
    """全量导出（主动权益 / 固收+ / 债券 / 混合）。"""
    with DorisConnector(ENV) as doris:
        logger.info(f"查询报告期 {REPORT_DATE}")
        df_equity  = _fetch(doris, _SQL_EQUITY,   _EQUITY_COLS)
        df_fi_plus = _fetch(doris, _SQL_FI_PLUS,  _FI_PLUS_COLS)
        df_bond    = _fetch(doris, _SQL_BOND,      _BOND_COLS)
        df_mix     = _fetch(doris, _SQL_MIX,       _MIX_COLS)

    logger.info(
        f"行数 — 主动权益:{len(df_equity)}  固收+:{len(df_fi_plus)}  "
        f"债券:{len(df_bond)}  混合:{len(df_mix)}"
    )

    _save(df_equity,              f'基金标签_主动权益_{REPORT_DATE}.xlsx')
    _save_fi_plus(df_fi_plus,     f'基金标签_固收+_{REPORT_DATE}.xlsx')
    _save(df_bond,                f'基金标签_债券_{REPORT_DATE}.xlsx')
    _save(df_mix,                 f'基金标签_混合_{REPORT_DATE}.xlsx')


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == 'fi':
        run_fi_plus()
    else:
        run()
