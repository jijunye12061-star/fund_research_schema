"""
客户交付 Excel 格式工具

使用场景区分：
  - 客户需求 → to_client_excel()：楷体 11pt、标题加粗居中、数字两位小数、备注页淡黄背景
  - 内部查数 → 直接 df.to_excel(path, index=False)，无需调用本模块

备注页规范：
  - 内容只写业务逻辑说明，不涉及任何字段名、表名、数据库术语
  - 两列：说明项 | 内容
"""
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── 样式常量（与用户确认的格式对齐，勿随意修改）─────────────────────────────
_FONT_NAME = '楷体'
_FONT_SIZE = 11
_NUM_FMT   = '0.00'
_NOTE_FILL = PatternFill('solid', fgColor='FFFFFFCC')   # 备注页淡黄背景
_NOTE_COL_WIDTHS = {'A': 15.73, 'B': 89.09}            # 备注页默认列宽


def _style_data_sheet(ws, col_widths: dict, num_cols: List[str]) -> None:
    """数据 sheet：楷体、标题行加粗居中、数字两位小数、列宽"""
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=(cell.row == 1))
            if cell.row == 1:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            if (cell.column_letter in num_cols
                    and cell.row > 1
                    and cell.value is not None):
                cell.number_format = _NUM_FMT
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width


def _style_notes_sheet(ws, col_widths: Optional[dict] = None) -> None:
    """备注 sheet：楷体、淡黄背景、标题居中、B列自动换行"""
    widths = col_widths or _NOTE_COL_WIDTHS
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=(cell.row == 1))
            cell.fill = _NOTE_FILL
            if cell.row == 1:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif cell.column_letter == 'B':
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def to_client_excel(
    path: str | Path,
    result_df: pd.DataFrame,
    notes_df: pd.DataFrame,
    col_widths: dict,
    num_cols: List[str],
    result_sheet: str = '结果',
    notes_col_widths: Optional[dict] = None,
) -> None:
    """
    输出客户交付格式 Excel，包含数据 sheet 和备注 sheet。

    参数：
        path           输出路径
        result_df      数据表
        notes_df       备注表，两列：['说明项', '内容']（仅含业务逻辑说明）
        col_widths     数据 sheet 列宽，如 {'A': 11, 'B': 42, ...}
        num_cols       需要保留两位小数的列字母列表，如 ['D', 'E', 'F']
        result_sheet   数据 sheet 名称（默认"结果"）
        notes_col_widths  备注 sheet 列宽，默认 {'A': 15.73, 'B': 89.09}
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name=result_sheet, index=False)
        notes_df.to_excel(writer, sheet_name='备注', index=False)
        _style_data_sheet(writer.sheets[result_sheet], col_widths, num_cols)
        _style_notes_sheet(writer.sheets['备注'], notes_col_widths)
